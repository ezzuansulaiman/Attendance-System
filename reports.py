"""
reports.py — Excel report builders for the attendance system.

build_internal_xlsx: Full grid with IC numbers, status colours, signature row.
build_external_xlsx: Summary counts only — for client/Air Selangor submission.
"""

import calendar
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

import db
from constants import DAY_ABBR_MS, REGIONS, STATUS_COLORS, STATUS_LABELS

# Excel fill colours (strip '#', openpyxl expects ARGB: FF + RRGGBB)
_FILLS = {
    code: PatternFill("solid", fgColor="FF" + color.lstrip("#"))
    for code, color in STATUS_COLORS.items()
}
_FILL_PH = PatternFill("solid", fgColor="FFFFF0C3")   # public holiday yellow
_FILL_HEADER = PatternFill("solid", fgColor="FF1F493C")  # brand dark green
_FILL_SUBHDR = PatternFill("solid", fgColor="FF2D6954")  # brand green
_FILL_DAYROW = PatternFill("solid", fgColor="FFE8EDE5")  # light grey-green
_FILL_TOTAL = PatternFill("solid", fgColor="FFDCE8DC")

_FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
_FONT_BOLD = Font(name="Calibri", bold=True, size=9)
_FONT_BOLD_SM = Font(name="Calibri", bold=True, size=8)
_FONT_SM = Font(name="Calibri", size=8)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)

_THIN = Side(style="thin", color="FF999999")
_THICK = Side(style="medium", color="FF444444")

_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_THICK = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)


def _set_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _BORDER_ALL


def _month_name_ms(month):
    names = ["Januari", "Februari", "Mac", "April", "Mei", "Jun",
             "Julai", "Ogos", "September", "Oktober", "November", "Disember"]
    return names[month - 1]


def build_internal_xlsx(region, year, month):
    """
    Internal report: full day-by-day grid with IC numbers, status cell colours,
    and a signature footer row. Suitable for internal filing.
    """
    employees, grid = db.get_month_grid(region, year, month)
    ph_dates = db.get_public_holiday_dates(year, month)
    num_days = calendar.monthrange(year, month)[1]
    region_name = REGIONS.get(region, region).upper()
    month_label = f"{_month_name_ms(month).upper()} {year}"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{region_name[:12]}_{month:02d}"
    ws.sheet_view.showGridLines = False

    # ── Column widths ──
    ws.column_dimensions["A"].width = 5   # No
    ws.column_dimensions["B"].width = 24  # Name
    ws.column_dimensions["C"].width = 16  # IC
    ws.column_dimensions["D"].width = 18  # Designation
    for col_i in range(5, 5 + num_days):
        ws.column_dimensions[get_column_letter(col_i)].width = 4.2
    # Total columns
    total_start = 5 + num_days
    for offset, w in enumerate([6, 5, 5, 5, 6, 8]):
        ws.column_dimensions[get_column_letter(total_start + offset)].width = w

    # ── Row 1: Main title ──
    total_cols = 4 + num_days + 6  # No+Name+IC+Desig + days + 6 totals
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1,
        value=f"LAPORAN KEHADIRAN PEKERJA — {region_name} — {month_label}")
    title_cell.font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=13)
    title_cell.fill = _FILL_HEADER
    title_cell.alignment = _ALIGN_C
    ws.row_dimensions[1].height = 26

    # ── Row 2: Column labels ──
    headers = ["No", "Nama Penuh", "No. IC", "Jawatan"]
    days_list = list(range(1, num_days + 1))
    day_abbr = [DAY_ABBR_MS[date(year, month, d).weekday()] for d in days_list]
    headers += [str(d) for d in days_list]
    headers += ["P", "AL", "MC", "EML", "OD/RD", "TANDATANGAN"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = _FONT_WHITE
        cell.fill = _FILL_SUBHDR
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL
    ws.row_dimensions[2].height = 22

    # ── Row 3: Day-of-week abbreviations ──
    ws.cell(row=3, column=1, value="").fill = _FILL_DAYROW
    ws.cell(row=3, column=2, value="").fill = _FILL_DAYROW
    ws.cell(row=3, column=3, value="").fill = _FILL_DAYROW
    ws.cell(row=3, column=4, value="").fill = _FILL_DAYROW
    for i, abbr in enumerate(day_abbr):
        col = 5 + i
        c = ws.cell(row=3, column=col, value=abbr)
        c.font = _FONT_BOLD_SM
        c.fill = _FILL_DAYROW
        c.alignment = _ALIGN_C
        c.border = _BORDER_ALL
    for offset in range(6):
        c = ws.cell(row=3, column=total_start + offset)
        c.fill = _FILL_DAYROW
        c.border = _BORDER_ALL
    ws.row_dimensions[3].height = 16

    # ── Data rows ──
    for row_i, emp in enumerate(employees, start=1):
        r = row_i + 3  # excel row
        ws.row_dimensions[r].height = 18

        ws.cell(row=r, column=1, value=row_i).font = _FONT_SM
        ws.cell(row=r, column=1).alignment = _ALIGN_C
        ws.cell(row=r, column=1).border = _BORDER_ALL

        ws.cell(row=r, column=2, value=emp["full_name"]).font = _FONT_BOLD_SM
        ws.cell(row=r, column=2).alignment = _ALIGN_L
        ws.cell(row=r, column=2).border = _BORDER_ALL

        ws.cell(row=r, column=3, value=emp.get("ic_number") or "").font = _FONT_SM
        ws.cell(row=r, column=3).alignment = _ALIGN_C
        ws.cell(row=r, column=3).border = _BORDER_ALL

        ws.cell(row=r, column=4, value=emp["designation"]).font = _FONT_SM
        ws.cell(row=r, column=4).alignment = _ALIGN_L
        ws.cell(row=r, column=4).border = _BORDER_ALL

        emp_grid = grid.get(emp["id"], {})
        counts = {s: 0 for s in ["P", "AL", "MC", "EML", "OD", "RD"]}

        for d in days_list:
            col = 4 + d
            date_str = f"{year}-{month:02d}-{d:02d}"
            status = emp_grid.get(d, "")
            c = ws.cell(row=r, column=col, value=status)
            c.font = _FONT_SM
            c.alignment = _ALIGN_C
            c.border = _BORDER_ALL
            if status in _FILLS:
                c.fill = _FILLS[status]
                counts[status] = counts.get(status, 0) + 1
            elif date_str in ph_dates and not status:
                c.fill = _FILL_PH

        # Totals
        od_rd = counts.get("OD", 0) + counts.get("RD", 0)
        for offset, val in enumerate([counts["P"], counts["AL"], counts["MC"],
                                      counts["EML"], od_rd, ""]):
            c = ws.cell(row=r, column=total_start + offset, value=val)
            c.font = _FONT_BOLD_SM
            c.alignment = _ALIGN_C
            c.border = _BORDER_ALL
            c.fill = _FILL_TOTAL

    # ── Legend row ──
    leg_row = len(employees) + 4
    ws.row_dimensions[leg_row].height = 14
    legend = "LEGENDA:  P=Hadir  OD=Hari Rehat  RD=Berehat  PH=Cuti Umum  AL=Cuti Tahunan  MC=Cuti Sakit  EML=Cuti Kecemasan"
    ws.merge_cells(start_row=leg_row, start_column=1,
                   end_row=leg_row, end_column=total_cols)
    lc = ws.cell(row=leg_row, column=1, value=legend)
    lc.font = Font(name="Calibri", italic=True, size=8, color="FF555555")
    lc.alignment = _ALIGN_L

    # ── Signature row ──
    sig_row = leg_row + 2
    ws.row_dimensions[sig_row].height = 36
    ws.merge_cells(start_row=sig_row, start_column=1,
                   end_row=sig_row, end_column=total_cols // 2)
    ws.merge_cells(start_row=sig_row, start_column=total_cols // 2 + 1,
                   end_row=sig_row, end_column=total_cols)
    ws.cell(row=sig_row, column=1,
        value="Disediakan oleh: _______________________________").font = _FONT_BOLD_SM
    ws.cell(row=sig_row, column=total_cols // 2 + 1,
        value="Disahkan oleh: _______________________________").font = _FONT_BOLD_SM

    # ── Page setup ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{get_column_letter(total_cols)}{sig_row}"
    ws.freeze_panes = "E4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_external_xlsx(region, year, month):
    """
    External report: summary counts only. No IC numbers, no day-by-day grid.
    For submission to Air Selangor / external parties.
    """
    summary = db.get_month_summary(region, year, month)
    num_days = calendar.monthrange(year, month)[1]
    region_name = REGIONS.get(region, region).upper()
    month_label = f"{_month_name_ms(month).upper()} {year}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    col_widths = [5, 28, 20, 14, 8, 8, 8, 10, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:I1")
    tc = ws.cell(row=1, column=1,
        value=f"ATTENDANCE SUMMARY — {region_name} — {month_label}")
    tc.font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=13)
    tc.fill = _FILL_HEADER
    tc.alignment = _ALIGN_C
    ws.row_dimensions[1].height = 26

    # Sub-header: printed date
    ws.merge_cells("A2:I2")
    dc = ws.cell(row=2, column=1,
        value=f"Tarikh Cetak: {date.today().strftime('%d %B %Y')}   |   Hari dalam Bulan: {num_days}")
    dc.font = Font(name="Calibri", italic=True, size=9, color="FF444444")
    dc.alignment = _ALIGN_L
    ws.row_dimensions[2].height = 16

    # Column headers
    col_hdrs = ["No", "Nama Penuh", "Jawatan",
                "Hari Hadir (P)", "AL", "MC", "EML", "Cuti Umum (PH)",
                "Rehat (OD+RD)"]
    for col, h in enumerate(col_hdrs, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _FONT_WHITE
        c.fill = _FILL_SUBHDR
        c.alignment = _ALIGN_C
        c.border = _BORDER_ALL
    ws.row_dimensions[3].height = 22

    # Data rows
    for i, emp in enumerate(summary, start=1):
        r = i + 3
        ws.row_dimensions[r].height = 18
        od_rd = emp.get("OD", 0) + emp.get("RD", 0)
        row_vals = [i, emp["full_name"], emp["designation"],
                    emp.get("P", 0), emp.get("AL", 0),
                    emp.get("MC", 0), emp.get("EML", 0),
                    emp.get("PH", 0), od_rd]
        for col, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _FONT_SM if col > 2 else _FONT_BOLD_SM
            c.alignment = _ALIGN_C if col != 2 else _ALIGN_L
            c.border = _BORDER_ALL

    # Totals row
    total_row = len(summary) + 4
    ws.row_dimensions[total_row].height = 20
    ws.cell(row=total_row, column=1, value="JUMLAH").font = _FONT_BOLD
    ws.cell(row=total_row, column=1).fill = _FILL_TOTAL
    ws.cell(row=total_row, column=1).alignment = _ALIGN_C
    for col in range(2, 10):
        c = ws.cell(row=total_row, column=col)
        c.fill = _FILL_TOTAL
        c.border = _BORDER_ALL
        if col >= 4:
            col_letter = get_column_letter(col)
            start = 4
            end = total_row - 1
            c.value = f"=SUM({col_letter}{start}:{col_letter}{end})"
            c.font = _FONT_BOLD
            c.alignment = _ALIGN_C

    # Signature
    sig_row = total_row + 3
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws.merge_cells(f"F{sig_row}:I{sig_row}")
    ws.cell(row=sig_row, column=1,
        value="Disediakan oleh: ____________________").font = _FONT_BOLD_SM
    ws.cell(row=sig_row, column=6,
        value="Disahkan oleh: ____________________").font = _FONT_BOLD_SM

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
